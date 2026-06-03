"""Build the CLKG data-collection template.

Multi-sheet xlsx: 填写说明(instructions) + 4 collection sheets + 词表(vocab).
Dropdowns reference DYNAMIC named ranges over 词表 columns (OFFSET/COUNTA), so
adding/removing a category at the bottom of a 词表 column auto-updates every
dropdown across all sheets ("关联表格").
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.comments import Comment

FONT = "Arial"
REQ_FILL = PatternFill("solid", fgColor="C00000")      # required header = dark red
OPT_FILL = PatternFill("solid", fgColor="1F4E78")      # optional header = dark blue
EXT_FILL = PatternFill("solid", fgColor="7030A0")      # project-extension header = purple
EX_FILL  = PatternFill("solid", fgColor="FFF2CC")      # example row = light yellow
VOCAB_FILL = PatternFill("solid", fgColor="548235")
HDR_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
EX_FONT  = Font(name=FONT, italic=True, color="7F6000", size=10)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

wb = Workbook()

# ---------------------------------------------------------------- 词表 (vocab)
vs = wb.active
vs.title = "词表"
VOCAB = {
    "entity_type": ["pl", "clu", "ac", "img", "doc", "ev"],
    "region":      ["mustang", "qiaopi", "xinjiang", "kashgar", "liangzhu"],
    "source_type": ["literature", "archive", "field_survey", "oral_history",
                    "gis_dataset", "web", "image_file", "other"],
    "CRS":         ["WGS84", "GCJ-02", "BD-09", "CGCS2000",
                    "EPSG:32644", "EPSG:32645", "unk"],
    "layer":       ["单体", "群体", "区域"],
    "heritage_lv": ["国家级", "自治区级", "省级", "市级", "县市级"],
    "confidence":  ["1.0", "0.9", "0.7", "0.5"],
    "actor_role":  ["person", "family", "organization"],
    "doc_type":    ["专著", "期刊论文", "档案", "方志", "碑刻铭文", "契约文书",
                    "信札(侨批)", "谱牒", "舆图", "报刊", "口述史转写", "图像图说", "其他"],
    "lang":        ["zh", "en", "zh-en双语", "其他"],
    "doc_status":  ["待验证", "已验证", "存疑", "废弃"],
}
VOCAB_NOTE = {
    "entity_type": "pl=单体地点 / clu=文化景观单元(群体·区域) / ac=人物机构 / img=图像音视频 / doc=文档(侨批·档案·口述史) / ev=事件",
    "region": "项目/区域代号，对应一个区域库。现有三库: mustang/qiaopi/xinjiang; kashgar/liangzhu 为规划。新项目可在本列底部新增。",
    "source_type": "literature=文献 archive=档案 field_survey=田野 oral_history=口述史 gis_dataset=GIS数据集 web=网络公开 image_file=图像 other=其他",
    "CRS": "坐标来源的坐标系！高德=GCJ-02，百度=BD-09，国外UTM-44N=EPSG:32644。绝不自己换算，ingest 统一转 WGS84。",
    "layer": "仅 Mustang 等需要时填。群体/区域→entity_type 记 clu；单体→pl。",
    "heritage_lv": "文物保护级别（如适用）。",
    "confidence": "对该条记录正确性的把握。亲见/原档=1.0，二手可靠=0.9，推断=0.7，存疑=0.5。",
    "actor_role": "person=个人 family=家族 organization=机构/商号。",
    "doc_type": "文献体裁：专著/期刊论文/档案/方志/碑刻铭文/契约文书/信札(侨批)/谱牒/舆图/报刊/口述史转写/图像图说/其他。",
    "lang": "文献主要语言：zh=中文 en=英文 zh-en双语 其他。",
    "doc_status": "数据状态(可选)：待验证/已验证/存疑/废弃。对应早期 design_legacy 的 Status 字段。",
}
cols = list(VOCAB.keys())
for j, key in enumerate(cols, start=1):
    c = vs.cell(row=1, column=j, value=key)
    c.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
    c.fill = VOCAB_FILL
    c.alignment = Alignment(horizontal="center")
    c.border = BORDER
    c.comment = Comment(VOCAB_NOTE[key], "SOP")
    vs.column_dimensions[c.column_letter].width = 15
    for i, val in enumerate(VOCAB[key], start=2):
        cc = vs.cell(row=i, column=j, value=val)
        cc.font = Font(name=FONT, size=10)
        cc.border = BORDER
vs.freeze_panes = "A2"
vs.sheet_properties.tabColor = "548235"

# Dynamic named ranges: dv_<key> = OFFSET('词表'!$X$2,0,0,COUNTA('词表'!$X:$X)-1,1)
# -> a dropdown bound to dv_<key> auto-extends when rows are appended to 词表.
for key in cols:
    j = cols.index(key) + 1
    L = vs.cell(row=1, column=j).column_letter
    formula = f"OFFSET('词表'!${L}$2,0,0,COUNTA('词表'!${L}:${L})-1,1)"
    wb.defined_names[f"dv_{key}"] = DefinedName(name=f"dv_{key}", attr_text=formula)

# ---------------------------------------------------------------- sheet builder
def build_sheet(name, tab, fields, example):
    """fields: list of (header, required, vocab_key_or_None, note, width[, ext_bool]).
    ext_bool=True marks a project-specific extension column (purple header)."""
    ws = wb.create_sheet(name)
    ws.sheet_properties.tabColor = tab
    for j, f in enumerate(fields, start=1):
        hdr, req, vkey, note, width = f[:5]
        ext = f[5] if len(f) > 5 else False
        c = ws.cell(row=1, column=j, value=hdr)
        c.font = HDR_FONT
        c.fill = EXT_FILL if ext else (REQ_FILL if req else OPT_FILL)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
        ws.column_dimensions[c.column_letter].width = width
        if note:
            c.comment = Comment(note, "SOP")
        if vkey:
            dv = DataValidation(type="list", formula1=f"dv_{vkey}", allow_blank=True)
            ws.add_data_validation(dv)
            dv.add(f"{c.column_letter}2:{c.column_letter}1000")
    for j, val in enumerate(example, start=1):
        c = ws.cell(row=2, column=j, value=val)
        c.font = EX_FONT
        c.fill = EX_FILL
        c.border = BORDER
    ws.cell(row=2, column=1).comment = Comment("← 金标准示例行，照着填，填完删掉本行。", "SOP")
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 42
    return ws

# ---------------------------------------------------------------- Place 地点
build_sheet("Place_地点", "2E75B6", [
    ("记录人*", True, None, "你的名字，便于追溯。", 10),
    ("采集日期*", True, None, "YYYY-MM-DD，记录当天日期。", 12),
    ("region*", True, "region", "项目/区域代号。", 11),
    ("entity_type*", True, "entity_type", "单体→pl；群体/区域→clu。", 11),
    ("natural_key*", True, None, "稳定唯一标识！同一地点永远用同一个 key（去重靠它）。优先用已有编号(遗产编码/照片文件夹名)，否则用 region-pl-有意义短码。", 18),
    ("hasName_名称*", True, None, "中文正式名称。", 16),
    ("hasType_类型", False, None, "如 庙宇/村落/墓葬。", 12),
    ("hasCategoryType_遗产类别", False, None, "如 古遗址/古建筑。", 14),
    ("hasEra_年代", False, None, "如 汉代/19c/清。", 11),
    ("hasLayer_层级", False, "layer", "Mustang 等：单体/群体/区域。", 11),
    ("hasHeritageLevel_保护级别", False, "heritage_lv", "如适用。", 12),
    ("lon_经度", False, None, "十进制度。务必同时填 CRS_坐标系。", 11),
    ("lat_纬度", False, None, "十进制度。", 11),
    ("CRS_坐标系", False, "CRS", "坐标来源的坐标系，必须如实声明！绝不自己换算。", 13),
    ("坐标来源", False, None, "如 高德地图/实测GPS/文献附图。", 14),
    ("hasPrefecture_省州", False, None, "", 12),
    ("hasCounty_县区", False, None, "", 12),
    ("hasTown_乡镇", False, None, "", 11),
    ("hasVillage_村", False, None, "", 11),
    ("hasAddress_地址", False, None, "完整地址。", 18),
    ("hasDescription_描述", False, None, "自由文本描述。", 24),
    ("hasSurveyDate_调查日期", False, None, "实地调查/资料对应日期。", 12),
    ("source_name_来源*", True, None, "出处！文献=Zotero引用键或题名页码；档案=馆藏号；网址=URL；田野=调查表编号。", 22),
    ("source_type_来源类型*", True, "source_type", "见词表。", 13),
    ("confidence_可信度", False, "confidence", "把握程度。", 11),
    ("asset_path_关联资产", False, None, "照片/扫描件的 NAS 路径或文件夹名。", 18),
    ("hasNotes_备注", False, None, "", 18),
], example=[
    "张三", "2026-06-01", "mustang", "pl", "mustang-pl-lomanthang", "洛满堂",
    "庙宇", "古建筑", "19c", "单体", "", "83.9567", "29.1845", "WGS84", "实测GPS",
    "甘达基省", "木斯塘县", "", "Lo Manthang", "古城北侧", "三层夯土建筑，保存完好。",
    "2024-05-18", "Mustang2024_field#012", "field_survey", "1.0",
    "/Volumes/clkg_data/mustang/photos/mustang-pl-lomanthang/", "需复查屋顶年代",
])

# ---------------------------------------------------------------- Document 文档（通用文本材料）
# 字段分三层(参考 design_legacy): 通用基础 → 文档核心(任何文献/档案/碑刻/口述史通用)
# → 项目专属扩展(紫色, 如侨批的寄收件人/金额, 仅相关项目填)。
build_sheet("Document_文档", "C55A11", [
    # —— 通用基础 ——
    ("记录人*", True, None, "", 10),
    ("采集日期*", True, None, "YYYY-MM-DD", 12),
    ("region*", True, "region", "", 11),
    ("entity_type*", True, "entity_type", "各类文本材料统一填 doc（文献/档案/方志/碑刻/契约/信札/口述史转写均可）。", 11),
    ("natural_key*", True, None, "稳定唯一标识。优先用已有编号(馆藏号/档号)；口述史用 受访者-日期 编号。", 18),
    # —— 文档核心(通用) ——
    ("hasTitle_标题*", True, None, "标题/题名；无题者据内容拟题；口述史可用 受访者口述(日期)。", 22),
    ("hasDocType_文献类型", False, "doc_type", "文献体裁，见词表(专著/档案/方志/碑刻/契约/信札…)。", 13),
    ("hasAuthor_作者创建者", False, None, "作者/编者/创建者；多个用、分隔。", 14),
    ("hasLanguage_语言", False, "lang", "主要语言。", 10),
    ("hasHoldingInstitution_收藏机构", False, None, "收藏/提供机构全称。", 16),
    ("hasCollectionNumber_馆藏号", False, None, "馆藏号/索书号/档号等。", 13),
    ("hasFormalDate_标准化日期", False, None, "成文/出版日期，公历，如 1929 或 1929-XX-XX。", 14),
    ("hasShownDate_原件题署日期", False, None, "原件上写的日期，如 民国18年/光绪三十年。", 14),
    ("hasFullText_全文转写", False, None, "正文/信文/口述全文转写（如已数字化）。", 30),
    ("hasAbstract_内容摘要", False, None, "一两句话的内容提要。", 24),
    ("relatedPlace_关联地点", False, None, "相关地点的 natural_key 或地名；并去 Place 表登记。", 16),
    ("relatedActor_关联人物", False, None, "相关人物/机构的 natural_key 或名；并去 Actor 表登记。", 16),
    ("hasReference_著录参考", False, None, "著录出处/参考文献/Zotero 键。", 16),
    ("hasRights_版权许可", False, None, "版权/使用许可范围。", 14),
    # —— 通用基础(尾) ——
    ("source_name_来源*", True, None, "出处：馆藏号/Zotero键/档号/录音文件名。", 22),
    ("source_type_来源类型*", True, "source_type", "文献=literature；档案=archive；口述史=oral_history。", 13),
    ("confidence_可信度", False, "confidence", "对该条记录正确性的把握。", 11),
    ("状态", False, "doc_status", "数据状态(可选)：待验证/已验证/存疑/废弃。", 10),
    ("asset_path_资产路径", False, None, "扫描件/录音/录像的 NAS 路径。", 20),
    ("hasNotes_备注", False, None, "", 18),
    # —— 项目专属扩展(侨批等批信类专用，其它项目留空) ——
    ("hasSender_寄件人(侨批)", False, None, "批信寄件人 natural_key/姓名；并去 Actor 表登记。", 14, True),
    ("hasRecipient_收件人(侨批)", False, None, "批信收件人。", 14, True),
    ("hasOriginPlace_寄出地(侨批)", False, None, "寄出地 natural_key/地名。", 14, True),
    ("hasDestinationPlace_寄达地(侨批)", False, None, "寄达地。", 14, True),
    ("hasReplyDate_回批日期(侨批)", False, None, "回批日期。", 12, True),
    ("hasCurrency_币种(侨批)", False, None, "如 墨西哥银元。", 12, True),
    ("hasAmount_金额(侨批)", False, None, "原始金额。", 11, True),
    ("hasPaidAmount_实付(侨批)", False, None, "实付金额。", 11, True),
    ("hasConvertedAmount_折算(侨批)", False, None, "标准化折算金额。", 12, True),
], example=[
    "李四", "2026-06-01", "qiaopi", "doc", "Q12345",
    "陈妙清寄潮安家书", "信札(侨批)", "陈妙清", "zh", "潮汕侨批馆", "Q12345",
    "1929-XX-XX", "民国18年", "父亲大人膝下敬禀者……", "寄银百元并报平安",
    "qiaopi-pl-chaoan", "qiaopi-ac-chenmiaoqing", "《潮汕侨批集成》第3辑", "馆藏，研究用途",
    "Q12345 / 潮汕侨批馆", "archive", "0.9", "待验证",
    "/Volumes/clkg_data/qiaopi/scans/Q12345.jpg", "字迹部分模糊",
    "qiaopi-ac-chenmiaoqing", "qiaopi-ac-chenmu", "qiaopi-pl-saigon", "qiaopi-pl-chaoan",
    "", "墨西哥银元", "100", "98", "550",
])

# ---------------------------------------------------------------- Actor 人物
build_sheet("Actor_人物", "7030A0", [
    ("记录人*", True, None, "", 10),
    ("采集日期*", True, None, "YYYY-MM-DD", 12),
    ("region*", True, "region", "", 11),
    ("entity_type*", True, "entity_type", "人物/家族/机构统一填 ac。", 11),
    ("natural_key*", True, None, "稳定唯一标识。被 Document 表的寄/收件人引用，必须一致！", 18),
    ("hasName_姓名*", True, None, "姓名/家族名/机构名。", 16),
    ("角色类型", False, "actor_role", "person/family/organization。", 12),
    ("hasDescription_说明", False, None, "身份、关系、生平等。", 26),
    ("source_name_来源*", True, None, "出处。", 20),
    ("source_type_来源类型*", True, "source_type", "", 13),
    ("confidence_可信度", False, "confidence", "", 11),
    ("hasNotes_备注", False, None, "", 18),
], example=[
    "李四", "2026-06-01", "qiaopi", "ac", "qiaopi-ac-chenmiaoqing", "陈妙清",
    "person", "潮安籍南洋华侨，1920s 寄批人。", "Q12345", "archive", "0.9", "",
])

# ---------------------------------------------------------------- Asset 资产
build_sheet("Asset_资产", "BF8F00", [
    ("记录人*", True, None, "", 10),
    ("采集日期*", True, None, "YYYY-MM-DD", 12),
    ("region*", True, "region", "", 11),
    ("entity_type*", True, "entity_type", "照片/录音/录像统一填 img。", 11),
    ("natural_key*", True, None, "稳定唯一标识（可用文件名去扩展名）。", 18),
    ("hasFileName_文件名*", True, None, "含扩展名，如 IMG_1234.jpg。", 16),
    ("hasFilePath_路径*", True, None, "NAS 完整路径 file://… 或相对 NAS_ROOT 的路径。", 24),
    ("hasFileType_类型", False, None, "JPEG/WAV/MP4 等。", 11),
    ("depicts_关联实体", False, None, "该资产描绘/对应的 Place 或 Document 的 natural_key。", 18),
    ("capturedAt_拍摄时间", False, None, "ISO 时间，如 2024-05-18T14:30。", 16),
    ("lon_经度", False, None, "EXIF GPS（通常 WGS84）。", 11),
    ("lat_纬度", False, None, "", 11),
    ("CRS_坐标系", False, "CRS", "EXIF 一般是 WGS84。", 12),
    ("hasAltitude_海拔", False, None, "米。", 11),
    ("source_name_来源*", True, None, "出处。", 20),
    ("source_type_来源类型*", True, "source_type", "照片填 image_file。", 13),
    ("hasNotes_备注", False, None, "", 18),
], example=[
    "张三", "2026-06-01", "mustang", "img", "IMG_1234", "IMG_1234.jpg",
    "/Volumes/clkg_data/mustang/photos/mustang-pl-lomanthang/IMG_1234.jpg",
    "JPEG", "mustang-pl-lomanthang", "2024-05-18T14:30", "83.9567", "29.1845",
    "WGS84", "3840", "相机原片", "image_file", "正立面",
])

# ---------------------------------------------------------------- 填写说明 (instructions)
ins = wb.create_sheet("填写说明")
ins.sheet_properties.tabColor = "FF0000"
ins.column_dimensions["A"].width = 4
ins.column_dimensions["B"].width = 110
ins.sheet_view.showGridLines = False
_r = [1]
def line(text="", kind="body"):
    r = _r[0]; _r[0] += 1
    c = ins.cell(row=r, column=2, value=text)
    if kind == "title":
        c.font = Font(name=FONT, bold=True, size=15, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F4E78")
        ins.row_dimensions[r].height = 26
    elif kind == "h":
        c.font = Font(name=FONT, bold=True, size=12, color="1F4E78")
        ins.row_dimensions[r].height = 22
    elif kind == "red":
        c.font = Font(name=FONT, bold=True, size=10, color="C00000")
    else:
        c.font = Font(name=FONT, size=10)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    return c

line("CLKG 采集模板 · 填写说明（请先读我）", "title")
line()
line("一句话：一行 = 一个对象；红色表头列必填；带下拉的列请用下拉选，不要手敲。", "red")
line()
line("第一步 · 选对表", "h")
line("· 一处地点 / 一片区域  → Place_地点（单体填 pl；群体或区域填 clu）")
line("· 一份文档（侨批/档案/口述史）→ Document_文档（entity_type 填 doc）")
line("· 一个人 / 家族 / 机构  → Actor_人物（填 ac）")
line("· 一张照片 / 录音 / 录像 → Asset_资产（填 img）")
line()
line("第二步 · 照示例填", "h")
line("每张表第 2 行是黄色「金标准示例行」，照着它的格式填。全部填完后，把示例行整行删掉再提交。")
line("鼠标悬停在表头上有批注，说明这一列怎么填。")
line()
line("表头颜色 = 字段分层", "h")
line("· 红色 = 必填。")
line("· 蓝色 = 通用 / 核心选填（任何项目都适用）。")
line("· 紫色 = 项目专属扩展字段（如 Document 表里侨批的寄收件人/币种/金额），", "red")
line("  只有相关项目才填，其他项目留空即可。")
line("（字段三层: 通用基础 → 类核心 → 项目扩展，沿用早期 design_legacy 的设计思路。）")
line()
line("三条红线（最容易错，记住这三条）", "h")
line("1) 坐标系绝不自己换算：高德截的填 GCJ-02，百度填 BD-09，手机/GPS/谷歌地球填 WGS84；", "red")
line("   原样填坐标 + 如实选 CRS_坐标系。换算由系统统一做。猜错坐标系=整批偏移几百米。")
line("2) natural_key 是对象的“身份证”：同一个对象永远填同一个 key；填前先查权威清单，", "red")
line("   别把别人已建过的对象又新建一遍（否则变“双胞胎”）。")
line("3) 每一行都要有来源：文献填 Zotero 引用键，档案填馆藏号，网络填 URL，田野填调查表号。", "red")
line()
line("第三步 · 提交前自查", "h")
line("□ 红色必填列都填了   □ 填了坐标的，CRS_坐标系一定选了   □ natural_key 查过、没重复")
line("□ 删掉了黄色示例行   □ 下拉列用选的、没手敲错别字（如 mustang 不要写成 Mustang）")
line()
line("提交方式", "h")
line("把填好的表格 + 相关资产，放进【收件口里你的子目录】，然后通知负责人。（具体路径会上给）")
line()
line("词表怎么维护（重要）", "h")
line("要增 / 删某个下拉选项（实体类型、地区、来源类型、坐标系等），只在「词表」对应列的“最底部”")
line("增删——所有表的下拉会自动更新。不要在采集表里手敲新值，也不要在词表中间留空行。")
line()
line("跨表关联", "h")
line("文档的寄/收件人、寄出/寄达地，请填对应对象的 natural_key；并记得去 Actor / Place 表")
line("给那个人 / 那个地各建一行。资产用 depicts_关联实体 填它对应地点/文档的 natural_key。")
line()
line("拿不准怎么办", "h")
line("留空 + 在 hasNotes_备注 里写明你的疑问，交审核时再定，不要瞎猜填。")
line()
line("━━━ 附 · 各表字段速查 ━━━", "title")
line()

line("【Place_地点】一处地点 / 一片区域（庙、村、遗址、聚落群…）", "h")
line("· 通用: 记录人 / 采集日期 / region / entity_type(单体填 pl，群体或区域填 clu) / natural_key")
line("· 核心: hasName 名称 · hasType 类型(庙宇/村落…) · hasCategoryType 遗产类别 · hasEra 年代 ·")
line("        hasLayer 层级(单体/群体/区域) · hasHeritageLevel 保护级别")
line("· 空间: lon 经度 / lat 纬度 + CRS 坐标系(填了坐标就必须选！) + 坐标来源")
line("· 行政: hasPrefecture 省州 / hasCounty 县区 / hasTown 乡镇 / hasVillage 村 / hasAddress 地址")
line("· 其它: hasDescription 描述 · hasSurveyDate 调查日期 · asset_path 关联资产(照片路径/文件夹)")
line("· 来源: source_name 来源 · source_type 来源类型 · confidence 可信度 · hasNotes 备注")
line()

line("【Document_文档】通用文本材料：文献/专著/档案/方志/碑刻/契约/谱牒/报刊/口述史/侨批…", "h")
line("· 通用: 记录人 / 采集日期 / region / entity_type(统一填 doc) / natural_key")
line("· 核心: hasTitle 标题 · hasDocType 文献类型 · hasAuthor 作者创建者 · hasLanguage 语言 ·")
line("        hasHoldingInstitution 收藏机构 · hasCollectionNumber 馆藏号 · hasFormalDate 标准化日期 ·")
line("        hasShownDate 原件题署日期 · hasFullText 全文转写 · hasAbstract 内容摘要 ·")
line("        relatedPlace 关联地点 · relatedActor 关联人物 · hasReference 著录参考 · hasRights 版权许可")
line("· 来源: source_name · source_type · confidence · 状态 · asset_path 资产路径 · hasNotes")
line("· 紫色扩展(仅侨批等批信类填，其它项目留空): 寄件人 / 收件人 / 寄出地 / 寄达地 / 回批日期 /", "red")
line("        币种 / 金额 / 实付 / 折算")
line()

line("【Actor_人物】一个人 / 家族 / 机构（寄批人、收批人、商号、作者…）", "h")
line("· 通用: 记录人 / 采集日期 / region / entity_type(统一填 ac) / natural_key")
line("· 核心: hasName 姓名(或家族名/机构名) · 角色类型(person/family/organization) · hasDescription 说明")
line("· 来源: source_name · source_type · confidence · hasNotes")
line("· 关系: 它的 natural_key 会被 Document 的寄/收件人、relatedActor 引用——务必前后一致。")
line()

line("【Asset_资产】一张照片 / 录音 / 录像等多模态文件", "h")
line("· 通用: 记录人 / 采集日期 / region / entity_type(统一填 img) / natural_key")
line("· 核心: hasFileName 文件名 · hasFilePath 路径(NAS) · hasFileType 类型 ·")
line("        depicts 关联实体(它拍的那个地点/文档的 natural_key) · capturedAt 拍摄时间 ·")
line("        lon / lat / CRS(EXIF 一般 WGS84) · hasAltitude 海拔")
line("· 来源: source_name · source_type · hasNotes")
line()

line("━━━ 附 · 下拉值与关联关系深度参考（招募会用） ━━━", "title")
line()

line("一、11 类下拉可选值（全部在「词表」sheet 集中维护）", "h")
line("· entity_type: pl=单体地点 · clu=文化景观单元(群体/区域) · ac=人物/家族/机构 · img=图像音视频 · doc=文档 · ev=事件")
line("· region: mustang / qiaopi / xinjiang / kashgar / liangzhu — 每项对应一个独立的 {region}_cl_kg 数据库，新项目在词表底部续行")
line("· source_type: literature(文献) / archive(档案) / field_survey(田野) / oral_history(口述史) / gis_dataset(GIS) / web(网络) / image_file(图像) / other")
line("· CRS: WGS84 / GCJ-02(高德) / BD-09(百度) / CGCS2000 / EPSG:32644(UTM 44N) / EPSG:32645(UTM 45N) / unk — 绝不自己换算！", "red")
line("· layer: 单体 / 群体 / 区域 — Mustang 类项目专用。单体→entity_type 填 pl；群体/区域→填 clu")
line("· heritage_lv: 国家级 / 自治区级 / 省级 / 市级 / 县市级")
line("· confidence: 1.0=亲见/原档 · 0.9=二手可靠 · 0.7=推断 · 0.5=存疑")
line("· actor_role: person(个人) / family(家族) / organization(机构/商号)")
line("· doc_type: 专著·期刊论文·档案·方志·碑刻铭文·契约文书·信札(侨批)·谱牒·舆图·报刊·口述史转写·图像图说·其他")
line("· lang: zh / en / zh-en双语 / 其他")
line("· doc_status: 待验证 / 已验证 / 存疑 / 废弃")
line()

line("二、各表使用了哪些下拉", "h")
line("· Place_地点:    region · entity_type(pl/clu) · source_type · CRS · confidence · layer · heritage_lv")
line("· Document_文档: region · entity_type(doc)    · source_type · confidence · doc_type · lang · doc_status")
line("· Actor_人物:    region · entity_type(ac)     · source_type · confidence · actor_role")
line("· Asset_资产:    region · entity_type(img)    · source_type · CRS")
line()

line("三、关联关系的四层机制", "h")
line()
line("【关联 1】跨表实体引用——靠 natural_key 串起来（最关键）", "red")
line("  · Document.hasSender_寄件人(侨批) / hasRecipient_收件人(侨批) → 引用 Actor.natural_key")
line("  · Document.hasOriginPlace_寄出地 / hasDestinationPlace_寄达地 → 引用 Place.natural_key")
line("  · Document.relatedPlace / relatedActor → 引用 Place / Actor 的 natural_key")
line("  · Asset.depicts_关联实体 → 引用它拍的那个 Place 或 Document 的 natural_key")
line("  · Place.asset_path_关联资产 → 资产 NAS 路径（弱关联，路径字符串）")
line("  填表规则：在 Document 写 hasSender=陈妙清 的同时，务必去 Actor 表给陈妙清单独建一行（用同一 natural_key）。", "red")
line()
line("【关联 2】词表 ↔ 采集表的动态绑定（'关联表格'）", "h")
line("  · 所有下拉绑定动态命名区域：dv_<key> = OFFSET('词表'!$X$2, 0, 0, COUNTA('词表'!$X:$X)-1, 1)")
line("  · 效果：在词表对应列【底部追加】新值——四张采集表的下拉自动扩展，无需重建模板。")
line("  · 禁区：不要在采集表里手敲词表外的值；不要在词表中间留空行；不要删除已被使用的值。", "red")
line()
line("【关联 3】entity_type 与表的硬约束", "h")
line("  · entity_type 下拉有 6 个值(pl/clu/ac/img/doc/ev)，但每张表只接受其中特定值：")
line("  · Place → 只能 pl 或 clu；Document → 只能 doc；Actor → 只能 ac；Asset → 只能 img")
line("  · ev(事件) 暂无专表，预留给未来扩展（田野调查事件、收藏事件等）。")
line()
line("【关联 4】同表内字段间约束（lon/lat ↔ CRS）", "h")
line("  · Place / Asset 表填了 lon_经度 或 lat_纬度，务必同时选 CRS_坐标系；")
line("  · Document 表无空间字段，不存在此约束。")
line()

# ---------------------------------------------------------------- order & save
order = ["填写说明", "Place_地点", "Document_文档", "Actor_人物", "Asset_资产", "词表"]
wb._sheets.sort(key=lambda s: order.index(s.title))
wb.active = 0
wb.save("/Users/lamue/clkg/data_collection/CLKG_采集模板.xlsx")
print("saved:", wb.sheetnames)
print("named ranges:", list(wb.defined_names))
