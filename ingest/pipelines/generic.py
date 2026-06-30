"""通用映射驱动入库器（L3 最后一公里）。

吃一份「映射 spec」（matching 建议 + 人工补的装配规则）+ xlsx，自动按规则装配成
实体 + 边并入库。把"任意新表 → 手写映射脚本"彻底降级成"填一份 spec"。

spec 结构（JSON）:
{
  "xlsx": "data/某表.xlsx",
  "sheet": "Sheet1",
  "region": "liangzhu",
  "source_type": "oral_history",
  "subject": {                         # 每行 = 一个主体实体
    "entity_type": "ev",
    "natural_key": {"synthetic_prefix": "liangzhu-oral"}   # 或 {"column": 5}
  },
  "columns": {                         # 列号(字符串) -> 规则
    "0":  {"kind": "value", "target": "hasFullText"},
    "1":  {"kind": "value", "target": "hasType"},
    "10": {"kind": "ref",   "target": "tookPlaceIn",  "ref_type": "pl", "ref_prefix": "liangzhu-place"},
    "13": {"kind": "ref",   "target": "carriedOutBy", "ref_type": "ac", "ref_prefix": "liangzhu-person"}
    # 省略的列 / kind=skip 一律忽略
  }
}

用法:
    python -m ingest.pipelines.generic <spec.json>           # 入库
    python -m ingest.pipelines.generic <spec.json> --dry-run # 只解析不写库
"""
from __future__ import annotations

import json
import sys
import uuid
from pathlib import Path
from typing import Any, Optional

import openpyxl

from .. import db, staging
from ..staging import StatementRow


def _clean(v: Any) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def rows_from_spec(spec: dict) -> list[StatementRow]:
    region = spec["region"]
    xlsx = Path(spec["xlsx"])
    sheet = spec.get("sheet")
    src_type = spec.get("source_type", "table_row")
    subj = spec["subject"]
    et = subj["entity_type"]
    nk_cfg = subj["natural_key"]
    cols = {int(k): v for k, v in spec["columns"].items()}

    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    next(it, None)  # header
    ev_base = f"file://{xlsx.resolve()}#{sheet or ws.title}"

    rows: list[StatementRow] = []
    ref_entities: dict[tuple[str, str], str] = {}  # (natural_key, type) -> name
    ordn = 0
    for r in it:
        if not any(c not in (None, "") for c in r):
            continue
        ordn += 1
        # 主体 natural_key
        if "column" in nk_cfg:
            nk = _clean(r[nk_cfg["column"]])
            if nk is None:
                continue
        else:
            nk = f"{nk_cfg['synthetic_prefix']}-{ordn:04d}"
        common = dict(
            ev_source_uri=f"{ev_base}&row={ordn + 1}", ev_source_type=src_type,
            ev_metadata={"spec": True}, ent_region=region, ent_type_abbr=et,
            ent_temporal="unk", ent_natural_key=nk)

        for cidx, cfg in cols.items():
            if cidx >= len(r):
                continue
            v = _clean(r[cidx])
            if v is None:
                continue
            kind = cfg.get("kind")
            if kind == "value":
                rows.append(StatementRow(**common, stmt_predicate=cfg["target"],
                                         stmt_value={"value": v}))
            elif kind == "ref":
                refnk = f"{cfg['ref_prefix']}-{v}"
                rtype = cfg.get("ref_type", "pl")
                rows.append(StatementRow(**common, stmt_predicate=cfg["target"],
                    stmt_value={"ref_natural_key": refnk, "ref_type_abbr": rtype,
                                "ref_region": region}))
                ref_entities[(refnk, rtype)] = v
            # kind in (skip / None / relation_meta) → 忽略

    # 给被引用实体补 hasName（让它们有标签）
    for (refnk, rtype), name in ref_entities.items():
        rows.append(StatementRow(
            ev_source_uri=f"curation://{region}/generic/entity",
            ev_source_type="entity_resolution", ev_metadata={}, ent_region=region,
            ent_type_abbr=rtype, ent_temporal="unk", ent_natural_key=refnk,
            stmt_predicate="hasName", stmt_value={"value": name}))
    wb.close()
    return rows


def run(spec: dict, dry_run: bool = False) -> dict:
    rows = rows_from_spec(spec)
    n_ent = len({(r.ent_type_abbr, r.ent_natural_key) for r in rows})
    n_edge = sum(1 for r in rows if "ref_natural_key" in r.stmt_value)
    print(f"[generic] 解析 → {len(rows)} 语句 / {n_ent} 实体 / {n_edge} 边")
    if dry_run:
        print("[generic] dry-run：不写库")
        return {"rows": len(rows), "entities": n_ent, "edges": n_edge}
    region = spec["region"]
    bid = uuid.uuid4()
    with db.connect(region) as conn:
        staging.write_rows(conn, bid, rows)
    with db.connect(region) as conn:
        rin, rok, rerr = staging.trigger_ingest_batch(conn, bid)
    print(f"[generic] ingest_batch: rows_in={rin} rows_ok={rok} rows_err={rerr}")
    return {"batch_id": str(bid), "rows_in": rin, "rows_ok": rok, "rows_err": rerr}


def main(argv: list[str]) -> int:
    if not argv:
        print("用法: python -m ingest.pipelines.generic <spec.json> [--dry-run]", file=sys.stderr)
        return 2
    spec = json.loads(Path(argv[0]).read_text(encoding="utf-8"))
    run(spec, dry_run="--dry-run" in argv)
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
