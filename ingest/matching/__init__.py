"""CLKG 自动列映射建议器（auto column mapping）。

给任意来源表的每一列，**建议**它对应的 CL-Onto 目标（谓词 / 角色 / 跨表引用），
靠"表头别名匹配 + 取值启发式"，不依赖 LLM。

定位：**建议，不自动应用**——映射错=数据错。输出供人工确认后，再喂给入库。
这是把"任意新表结构 → 手工写映射脚本"那一步，降级成"审一份建议表"。

用法:
    from ingest.matching import suggest_for_xlsx
    proposals = suggest_for_xlsx("data/某表.xlsx", "Sheet1")
命令行:
    python -m ingest.matching data/某表.xlsx Sheet1
"""
from __future__ import annotations

import re
from typing import Any, Optional

from ..collection_schema import ALL_SHEETS, EXTENSION_SHEETS, VOCAB

# ── 目标别名表：target -> [表头别名]（中英 + 常见变体）。会再叠加 schema 表头/中文名 ──
_SYNONYMS: dict[str, list[str]] = {
    # 角色
    "NATURAL_KEY": ["natural_key", "编号", "id", "标识", "唯一标识", "key", "序号"],
    "ENTITY_TYPE": ["entity_type", "实体类型"],
    "REGION":      ["region", "区域", "项目代号"],
    "LON":         ["经度", "lon", "longitude"],
    "LAT":         ["纬度", "lat", "latitude"],
    "CRS":         ["坐标系", "crs", "srid", "投影"],
    "SOURCE_NAME": ["来源", "出处", "source", "资料来源"],
    "META":        ["记录人", "采集日期", "校对人", "录入人", "著录人", "整理人"],
    # 谓词（字面值）
    "hasName":        ["名称", "名字", "name", "命名"],
    "hasTitle":       ["标题", "题名", "篇名", "title"],
    "hasType":        ["类型", "大事件类型", "事件类型", "种类", "type", "category", "类别"],
    "hasEra":         ["年代", "时代", "朝代", "era", "period", "历史时期"],
    "hasFullText":    ["全文", "正文", "原文", "句子原文", "内容", "fulltext", "text"],
    "hasDescription": ["描述", "说明", "简介", "详细事件", "事件描述", "description"],
    "hasAuthor":      ["作者", "创建者", "编者", "author", "creator"],
    "hasAddress":     ["地址", "住址", "address", "具体地点"],
    "hasShownDate":   ["日期", "整理时间", "具体年份", "date"],
    "hasTimeSpan":    ["时间段", "时段", "time-span", "timespan", "起止时间", "time span"],
    "hasObject":      ["物品", "人造物", "object", "artifact", "man-made"],
    "hasObjectType":  ["人造物分类", "物品分类", "object type"],
    "hasReference":   ["参考", "参考文献", "著录", "reference"],
    "hasProfession":  ["专业", "职业", "profession"],
    "hasLanguage":    ["语言", "language", "lang"],
    # 跨表引用
    "REF_PL": ["地点", "地名", "place", "place name", "地理位置", "发生地"],
    "REF_AC": ["人物", "人员", "person", "actor", "机构", "group", "组织", "口述者", "采访人", "发起人"],
    "REF_DOC": ["文档", "文献", "document", "关联文档"],
}

# 表头含这些 → 判为 "CRM 关系/属性标注列"，跳过（标边方向，非数据值）
_RELATION_HINTS = re.compile(
    r"关系|consists|took\s*place|carried\s*out|has\s*type|has\s*time|"
    r"\bP\d+\b|包含|发生在|执行者|有类型|有时间|有执行")

_CODE = re.compile(r"[EeＥ]\d+|[Pp]\d+")     # CRM 代码 E48/P7
_PAREN = re.compile(r"[（）()]")


def _norm(h: str) -> str:
    s = _CODE.sub(" ", str(h))
    s = _PAREN.sub(" ", s)
    return re.sub(r"\s+", " ", s).strip().lower()


def _kind_of(t: str) -> str:
    if t.startswith("REF_"):
        return "ref"
    if t in ("NATURAL_KEY", "ENTITY_TYPE", "REGION", "LON", "LAT", "CRS",
             "SOURCE_NAME", "META"):
        return "role"
    return "predicate"


def _alias_index() -> dict[str, tuple[str, str]]:
    idx: dict[str, tuple[str, str]] = {}

    def add(alias: str, target: str):
        a = _norm(alias)
        if a and a not in idx:
            idx[a] = (target, _kind_of(target))

    for tgt, aliases in _SYNONYMS.items():
        for a in aliases:
            add(a, tgt)
    for sd in ALL_SHEETS + EXTENSION_SHEETS:
        for col in sd.columns:
            tgt = col.predicate or col.role.name
            for a in (col.header, col.chinese_label):
                add(a, tgt)
    return idx


_IDX = _alias_index()


def _nums(vals: list[Any]) -> tuple[float, list[float]]:
    ns = []
    for v in vals:
        try:
            ns.append(float(str(v).strip()))
        except (TypeError, ValueError):
            pass
    return (len(ns) / len(vals) if vals else 0.0), ns


def _suggest_one(header: str, samples: list[Any]) -> dict:
    nh = _norm(header)
    sv = [s for s in samples if s not in (None, "")][:20]

    if _RELATION_HINTS.search(str(header)):
        return {"target": None, "kind": "relation_meta", "score": 0.9,
                "reason": "CRM 关系/属性标注列（标边方向，非数据）"}

    best = None  # (header_cover, alias_len, (target,kind))
    for alias, tk in _IDX.items():
        if alias and alias in nh:
            cover = len(alias) / max(len(nh), 1)
            cand = (cover, len(alias), tk)
            if best is None or cand[:2] > best[:2]:
                best = cand
    if best and best[0] >= 0.34:
        tgt, kind = best[2]
        return {"target": tgt, "kind": kind,
                "score": round(min(0.99, 0.5 + best[0] / 2), 2), "reason": "表头匹配别名"}

    frac, nums = _nums(sv)
    if sv and {str(x).strip() for x in sv} <= set(VOCAB["entity_type"]):
        return {"target": "ENTITY_TYPE", "kind": "role", "score": 0.8, "reason": "取值全是实体类型词"}
    if frac >= 0.7 and nums:
        if all(-180 <= n <= 180 for n in nums) and any(abs(n) > 90 for n in nums):
            return {"target": "LON", "kind": "role", "score": 0.6, "reason": "数值含|值|>90 像经度"}
        if all(-90 <= n <= 90 for n in nums):
            return {"target": "LAT", "kind": "role", "score": 0.45, "reason": "数值±90 像纬度"}
    if sv and sum(bool(re.search(r"\d{3,4}\s*年|\d{4}-\d", str(x))) for x in sv) >= len(sv) * 0.6:
        return {"target": "hasShownDate", "kind": "predicate", "score": 0.55, "reason": "取值像年份/日期"}

    return {"target": "hasNotes", "kind": "predicate", "score": 0.2,
            "reason": "未匹配——建议人工指定（暂落 hasNotes）"}


def suggest_mapping(headers: list[str], sample_rows: list[list[Any]]) -> list[dict]:
    cols = list(zip(*sample_rows)) if sample_rows else [() for _ in headers]
    out = []
    for i, h in enumerate(headers):
        samples = list(cols[i]) if i < len(cols) else []
        out.append({"col": i, "header": h, **_suggest_one(h or "", samples)})
    return out


def suggest_for_xlsx(path: str, sheet: Optional[str] = None, n_samples: int = 15) -> list[dict]:
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    headers = list(next(it, ()))
    samples = []
    for r in it:
        samples.append(list(r))
        if len(samples) >= n_samples:
            break
    wb.close()
    return suggest_mapping(headers, samples)
