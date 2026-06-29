"""CLKG standard template connector.

Reads a filled CLKG_采集模板.xlsx and emits StatementRow lists.
One row in the template → one entity (with many statements).
Cross-entity references (Document → Actor, Document → Place, Asset → Place/Doc)
are emitted as ref_natural_key statements — ingest_batch() resolves them.

See ingest/collection_schema.py for the column definitions consumed here.
"""
from __future__ import annotations

import logging
from collections import Counter
from pathlib import Path
from typing import Any, Optional

import openpyxl

from ..collection_schema import (
    ALL_SHEETS,
    EXTENSION_SHEETS,
    SHEETS_BY_NAME,
    ColumnDef,
    ColumnRole,
    SheetDescriptor,
    column_map,
)
from ..staging import StatementRow

log = logging.getLogger(__name__)

# ---- known example-row fill color (EX_FILL from build_template.py) --------
# Stored as the 6-hex RGB. openpyxl reads fills back as 8-hex ARGB
# (e.g. "00FFF2CC" / "FFFFF2CC"), so always compare the trailing 6 hex.
_EXAMPLE_FILL_RGB = "FFF2CC"

# ---- geometry SRID mapping -------------------------------------------------
_CRS_TO_SRID: dict[str, int] = {
    "WGS84": 4326,
    "CGCS2000": 4490,
    "EPSG:32644": 32644,
    "EPSG:32645": 32645,
}
_CRS_NEEDS_CONVERSION = {"GCJ-02", "BD-09"}
_CRS_UNKNOWN = "unk"


# ═══════════════════════════════════════════════════════════════════════════
# helpers
# ═══════════════════════════════════════════════════════════════════════════

def _clean(v: Any) -> Optional[str]:
    """Normalise a cell value: None / NaN / empty → None."""
    if v is None:
        return None
    if isinstance(v, float):
        try:
            import math
            if math.isnan(v):
                return None
        except (TypeError, ImportError):
            pass
    s = str(v).strip()
    return s if s and s.lower() not in ("nan", "none") else None


def _float_or_none(v: Any) -> Optional[float]:
    """Try to parse a cell as float; return None on failure."""
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def _example_fill_match(cell) -> bool:
    """True if a cell carries the gold-standard example fill (light yellow)."""
    try:
        if cell.fill and cell.fill.fgColor:
            rgb = cell.fill.fgColor.rgb
            # openpyxl returns 8-hex ARGB ("00FFF2CC"/"FFFFF2CC") or a non-str
            # theme object — compare the trailing 6 hex (RGB), case-insensitive.
            if isinstance(rgb, str) and rgb[-6:].upper() == _EXAMPLE_FILL_RGB:
                return True
    except Exception:
        pass
    return False


def _is_example_row(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    row_idx: int,
    sheet_name: str,
    nk_col: int = 5,
) -> bool:
    """Detect the gold-standard example row.

    Content/format-based detection — robust to row insertion *and* deletion:
      1. Fill colour == EX_FILL (light yellow) on the col-1 cell
      2. natural_key matches the sheet's known example key

    We deliberately do NOT treat "row 2" as the example by position alone.
    The template instructions tell collectors to delete the example row before
    submitting, which shifts their first real record up to row 2 — skipping it
    by position would silently drop that record.
    """
    if _example_fill_match(ws.cell(row=row_idx, column=1)):
        return True

    sd = SHEETS_BY_NAME.get(sheet_name)
    if sd and sd.example_natural_key:
        nk = _clean(ws.cell(row=row_idx, column=nk_col).value)
        if nk and nk == sd.example_natural_key:
            return True

    return False


def _is_empty_row(ws, row_idx: int, critical_cols: list[int]) -> bool:
    """True if every critical column is empty/null."""
    return all(
        _clean(ws.cell(row=row_idx, column=c).value) is None
        for c in critical_cols
    )


def _detect_region(ws, sd: SheetDescriptor, max_scan: int = 50) -> Optional[str]:
    """Find the most common region value in non-example data rows."""
    cm = column_map(sd)
    region_col = next((c.index for c in sd.columns if c.role == ColumnRole.REGION), None)
    if region_col is None:
        # Extension sheets (flat event tables) carry no region column — fall back
        # to the descriptor's fixed region.
        return sd.default_region
    counter: Counter[str] = Counter()
    for row_idx in range(2, min(ws.max_row or 0, max_scan + 2) + 1):
        if _is_example_row(ws, row_idx, sd.sheet_name):
            continue
        v = _clean(ws.cell(row=row_idx, column=region_col).value)
        if v:
            counter[v] += 1
    if not counter:
        return None
    top, count = counter.most_common(1)[0]
    if len(counter) > 1:
        second = counter.most_common(2)[1][1]
        if second == count:
            log.warning("Ambiguous region in %s: %s — using %s", sd.sheet_name, dict(counter), top)
    return top


def _guess_ref_type(natural_key: str) -> str:
    """Heuristic: inspect natural_key for type prefixes like -pl- / -ac- / -doc-."""
    nk = natural_key.lower()
    if "-pl-" in nk:   return "pl"
    if "-clu-" in nk:  return "clu"
    if "-ac-" in nk:   return "ac"
    if "-doc-" in nk:  return "doc"
    if "-img-" in nk:  return "img"
    return "pl"  # safest default; places are most common referent


# ═══════════════════════════════════════════════════════════════════════════
# row → StatementRow emission
# ═══════════════════════════════════════════════════════════════════════════

def _emit_row_statements(
    ws,
    row_idx: int,
    sd: SheetDescriptor,
    region: str,
    ev_uri_base: str,
    synthetic_ordinal: int = 0,
) -> list[StatementRow]:
    """Convert one template row into a list of StatementRows.

    Accumulates geometry (lon/lat/crs) and emits a single locatedAt statement
    at the end.  ``synthetic_ordinal`` is the 1-based position of this row among
    the sheet's data rows; used to mint a natural_key for extension sheets that
    lack a NATURAL_KEY column (e.g. 军垦_事件 → junken-evt-0001).
    """
    cm = column_map(sd)
    rows: list[StatementRow] = []

    # -- natural_key: from column, else synthesised by row order --------------
    nk_cols = [c for c in sd.columns if c.role == ColumnRole.NATURAL_KEY]
    if nk_cols:
        nk = _clean(ws.cell(row=row_idx, column=nk_cols[0].index).value)
    elif sd.synthetic_key_prefix:
        nk = f"{sd.synthetic_key_prefix}-{synthetic_ordinal:04d}"
    else:
        return rows
    if nk is None:
        return rows

    # entity_type: from column, else the descriptor's fixed default
    et_col = next((c for c in sd.columns if c.role == ColumnRole.ENTITY_TYPE), None)
    et_cell = _clean(ws.cell(row=row_idx, column=et_col.index).value) if et_col else None
    type_abbr = (et_cell or sd.default_entity_type
                 or (sd.entity_types[0] if sd.entity_types else "pl"))

    # temporal: try hasEra column first, else "unk"
    era_col = next((c for c in sd.columns if c.predicate == "hasEra"), None)
    temporal = "unk"
    if era_col:
        tv = _clean(ws.cell(row=row_idx, column=era_col.index).value)
        if tv:
            temporal = tv

    # -- evidence metadata ----------------------------------------------------
    ev_meta: dict[str, Any] = {
        "sheet":      sd.sheet_name,
        "row":        row_idx,
        "collector":  None,
        "recording_date": None,
    }
    source_name_val: Optional[str] = None
    source_type_val: Optional[str] = "template_row"
    confidence_val: Optional[float] = None

    for col in sd.columns:
        raw = ws.cell(row=row_idx, column=col.index).value
        v = _clean(raw)
        match col.role:
            case ColumnRole.META:
                if v:
                    ev_meta[col.chinese_label] = v
                    if "记录人" in col.header:
                        ev_meta["collector"] = v
                    if "采集日期" in col.header:
                        ev_meta["recording_date"] = v
            case ColumnRole.SOURCE_NAME:
                if v:
                    source_name_val = v
            case ColumnRole.SOURCE_TYPE:
                if v:
                    source_type_val = v
            case ColumnRole.CONFIDENCE:
                fv = _float_or_none(raw)
                if fv is not None:
                    confidence_val = fv
            case _:
                pass

    # Build evidence URI
    ev_uri = f"{ev_uri_base}#{sd.sheet_name}&row={row_idx}"

    # common kwargs for every StatementRow from this entity
    common = dict(
        ev_source_uri=ev_uri,
        ev_source_type=source_type_val or "template_row",
        ev_metadata=ev_meta,
        ent_region=region,
        ent_type_abbr=type_abbr,
        ent_temporal=temporal,
        ent_natural_key=nk,
    )

    # -- geometry accumulator -------------------------------------------------
    lon: Optional[float] = None
    lat: Optional[float] = None
    crs_value: Optional[str] = None

    # -- emit value & ref columns ---------------------------------------------
    for col in sd.columns:
        raw = ws.cell(row=row_idx, column=col.index).value
        v = _clean(raw)
        if v is None:
            continue

        match col.role:
            # ---- self-value statements ---------------------------------
            case ColumnRole.VALUE:
                rows.append(StatementRow(
                    **common,
                    stmt_predicate=col.predicate,
                    stmt_value={"value": v},
                ))

            # ---- cross-entity references --------------------------------
            case ColumnRole.REF_PL:
                rows.append(StatementRow(
                    **common,
                    stmt_predicate=col.predicate,
                    stmt_value={
                        "ref_natural_key":  v,
                        "ref_type_abbr":    "pl",
                        "ref_region":       region,
                    },
                ))
            case ColumnRole.REF_AC:
                rows.append(StatementRow(
                    **common,
                    stmt_predicate=col.predicate,
                    stmt_value={
                        "ref_natural_key":  v,
                        "ref_type_abbr":    "ac",
                        "ref_region":       region,
                    },
                ))
            case ColumnRole.REF_DOC:
                rows.append(StatementRow(
                    **common,
                    stmt_predicate=col.predicate,
                    stmt_value={
                        "ref_natural_key":  v,
                        "ref_type_abbr":    "doc",
                        "ref_region":       region,
                    },
                ))
            case ColumnRole.REF_ANY:
                guessed = _guess_ref_type(v)
                log.debug("REF_ANY %s → guessed type %s", v, guessed)
                rows.append(StatementRow(
                    **common,
                    stmt_predicate=col.predicate,
                    stmt_value={
                        "ref_natural_key":  v,
                        "ref_type_abbr":    guessed,
                        "ref_region":       region,
                    },
                ))

            # ---- asset path (weak ref) ----------------------------------
            case ColumnRole.ASSET_PATH:
                rows.append(StatementRow(
                    **common,
                    stmt_predicate=col.predicate,
                    stmt_value={"value": v},
                ))

            # ---- geometry (accumulated) ---------------------------------
            case ColumnRole.LON:
                lon = _float_or_none(raw)
            case ColumnRole.LAT:
                lat = _float_or_none(raw)
            case ColumnRole.CRS:
                crs_value = v

            # ---- roles consumed above (META, SOURCE, etc.) → no-op ------
            case _:
                pass

    # -- emit geometry if present ---------------------------------------------
    if lon is not None and lat is not None:
        if not (-180 <= lon <= 180 and -90 <= lat <= 90):
            log.warning("%s row %d: coords out of range (%.6f, %.6f) — skipped geometry",
                        sd.sheet_name, row_idx, lon, lat)
        else:
            geom_value: dict[str, Any] = {
                "wkt":        f"POINT({lon} {lat})",
                "source":     "template_xlsx",
                "source_crs": crs_value or "unspecified",
            }

            if crs_value in _CRS_NEEDS_CONVERSION:
                log.warning(
                    "%s row %d: CRS=%s — stored as-is (offset 50–500m possible). "
                    "GCJ-02/BD-09 coordinate conversion not yet implemented; "
                    "reprocess when coords module is ready.",
                    sd.sheet_name, row_idx, crs_value,
                )
                # Store as WGS84-ish with a flag so rows are findable later
                geom_value["srid"] = 4326
                geom_value["needs_conversion"] = crs_value

            elif crs_value and crs_value in _CRS_TO_SRID:
                geom_value["srid"] = _CRS_TO_SRID[crs_value]

            elif crs_value and crs_value.upper().startswith("EPSG:"):
                try:
                    geom_value["srid"] = int(crs_value.upper().split(":")[1])
                except (ValueError, IndexError):
                    log.warning("%s row %d: unparseable EPSG '%s' — using 4326",
                                sd.sheet_name, row_idx, crs_value)
                    geom_value["srid"] = 4326

            elif crs_value == _CRS_UNKNOWN:
                log.warning("%s row %d: CRS=unk — skipping geometry", sd.sheet_name, row_idx)
            else:
                # No CRS declared → assume WGS84
                geom_value["srid"] = 4326
                if crs_value is None:
                    pass  # quiet default; most EXIF coords are WGS84 anyway
                else:
                    log.info("%s row %d: unrecognized CRS '%s' — assuming 4326",
                             sd.sheet_name, row_idx, crs_value)

            if "srid" in geom_value:
                rows.append(StatementRow(
                    **common,
                    stmt_predicate="locatedAt",
                    stmt_value=geom_value,
                ))

    # -- apply row-level confidence into each statement's value JSONB ----------
    #    ingest_batch() reads stmt_value->>'confidence' (03_ingest_batch.sql),
    #    defaulting to 1.0 when absent. StatementRow has no confidence field, so
    #    the value must ride inside stmt_value to reach entity_statement.confidence.
    if confidence_val is not None:
        for r in rows:
            r.stmt_value["confidence"] = confidence_val

    return rows


# ═══════════════════════════════════════════════════════════════════════════
# per-sheet parser
# ═══════════════════════════════════════════════════════════════════════════

def _parse_sheet(
    ws,
    sd: SheetDescriptor,
    *,
    region: str,
    ev_uri_base: str,
    max_rows: Optional[int] = None,
    skip_example_row: bool = True,
) -> list[StatementRow]:
    """Parse one template sheet → list[StatementRow]."""
    rows: list[StatementRow] = []
    nk_cols = [c.index for c in sd.columns if c.role == ColumnRole.NATURAL_KEY]
    critical_cols = nk_cols + [
        c.index for c in sd.columns
        if c.role in (ColumnRole.VALUE, ColumnRole.ENTITY_TYPE)
        and c.required
    ]

    synth_ordinal = 0  # 1-based position among data rows; feeds synthetic keys
    for row_idx in range(2, ws.max_row + 1):  # row 1 = header
        if skip_example_row and _is_example_row(ws, row_idx, sd.sheet_name):
            continue

        if _is_empty_row(ws, row_idx, critical_cols):
            continue

        synth_ordinal += 1
        entity_rows = _emit_row_statements(
            ws, row_idx, sd, region, ev_uri_base,
            synthetic_ordinal=synth_ordinal,
        )
        rows.extend(entity_rows)

        if max_rows and len({r.ent_natural_key for r in rows}) >= max_rows:
            break

    log.info("%s: %d rows → %d statements", sd.sheet_name,
             len({r.ent_natural_key for r in rows}), len(rows))
    return rows


# ═══════════════════════════════════════════════════════════════════════════
# public entry point
# ═══════════════════════════════════════════════════════════════════════════

def ingest_template_xlsx(
    xlsx_path: Path,
    *,
    region: Optional[str] = None,
    max_rows: Optional[int] = None,
    skip_example_row: bool = True,
    strict_crs: bool = False,
    sheet_names: Optional[list[str]] = None,
) -> list[StatementRow]:
    """Read a CLKG standard template xlsx and emit StatementRows.

    Parameters
    ----------
    xlsx_path : Path
        Path to the filled template xlsx.
    region : str or None
        Override auto-detected region. Auto-detects if None.
    max_rows : int or None
        Per-sheet row limit (entity count, not statement count).  Useful for
        testing with a subset of rows.
    skip_example_row : bool
        Skip the gold-standard example row (row 2).  Default True.
    strict_crs : bool
        If True, raise on GCJ-02/BD-09 coords instead of logging a warning.
        (Not yet enforced — see §10.4 of the technical manual.)
    sheet_names : list[str] or None
        Which sheets to parse (e.g. ["Place_地点"]).  Default: all four.

    Returns
    -------
    list[StatementRow]
        Ready for staging.stg_ingest via staging.write_rows().
    """
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Template not found: {xlsx_path}")

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ev_uri_base = f"file://{xlsx_path.resolve()}"

    # Default target = the four generic sheets, plus any project-extension
    # sheets (军垦_*) that actually exist in this workbook.
    if sheet_names:
        target_sheets = sheet_names
    else:
        target_sheets = [sd.sheet_name for sd in ALL_SHEETS]
        target_sheets += [sd.sheet_name for sd in EXTENSION_SHEETS
                          if sd.sheet_name in wb.sheetnames]

    # --- region detection ----------------------------------------------------
    if region is None:
        for sn in target_sheets:
            sd = SHEETS_BY_NAME.get(sn)
            if sd and sn in wb.sheetnames:
                region = _detect_region(wb[sn], sd)
                if region:
                    break
    if region is None:
        raise ValueError(
            "Could not auto-detect region from template. "
            "Pass --region explicitly or ensure the template has at least one "
            "data row with a valid region value."
        )
    log.info("Using region=%s", region)

    # --- parse each sheet ----------------------------------------------------
    all_rows: list[StatementRow] = []
    for sn in target_sheets:
        sd = SHEETS_BY_NAME.get(sn)
        if sd is None:
            log.warning("Unknown sheet '%s' — skipped", sn)
            continue
        if sn not in wb.sheetnames:
            log.warning("Sheet '%s' not found in xlsx — skipped", sn)
            continue
        ws = wb[sn]
        sheet_rows = _parse_sheet(
            ws, sd, region=region, ev_uri_base=ev_uri_base,
            max_rows=max_rows, skip_example_row=skip_example_row,
        )
        all_rows.extend(sheet_rows)

    if strict_crs:
        # Future: fail if any unconverted CRS detected
        pass

    log.info("Total: %d statements from %d entities across %d sheets",
             len(all_rows),
             len({r.ent_natural_key for r in all_rows}),
             len(target_sheets))

    wb.close()
    return all_rows
