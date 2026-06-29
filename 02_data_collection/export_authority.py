"""Export the authority list of existing entities (natural_key) from the three
region databases, so collectors can look up & reuse keys instead of recreating
entities. Output: data_collection/权威实体清单.xlsx (one sheet per region) + per-region CSV.
"""
import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from ingest.db import connect

REGIONS = ["mustang", "qiaopi", "xinjiang"]
OUT = Path(__file__).resolve().parent
FONT = "Arial"

QUERY = """
SELECT r.type_abbr        AS entity_type,
       r.natural_key      AS natural_key,
       e.label_zh         AS label_zh,
       r.pid              AS pid
FROM pid_registry r
LEFT JOIN conceptual_entity e ON e.pid = r.pid
ORDER BY r.type_abbr, r.natural_key;
"""

wb = Workbook()
wb.remove(wb.active)
HDR_FILL = PatternFill("solid", fgColor="1F4E78")
HDR_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADERS = ["entity_type", "natural_key", "label_zh", "pid"]

summary = {}
for region in REGIONS:
    rows = []
    try:
        with connect(region) as conn, conn.cursor() as cur:
            cur.execute(QUERY)
            rows = cur.fetchall()
    except Exception as e:
        print(f"[WARN] {region}: {e}")
        continue

    # per-region CSV
    with open(OUT / f"权威清单_{region}.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(HEADERS)
        w.writerows(rows)

    # per-region sheet
    ws = wb.create_sheet(region)
    for j, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.font = HDR_FONT; c.fill = HDR_FILL; c.border = BORDER
        c.alignment = Alignment(horizontal="center")
    widths = [12, 32, 28, 34]
    for j, wd in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=j).column_letter].width = wd
    for i, row in enumerate(rows, start=2):
        for j, val in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=val)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:D{max(1, len(rows) + 1)}"

    # per-type counts
    by_type = {}
    for et, *_ in rows:
        by_type[et] = by_type.get(et, 0) + 1
    summary[region] = (len(rows), by_type)

wb.save(OUT / "权威实体清单.xlsx")

print("=== 权威清单导出完成 ===")
for region in REGIONS:
    if region in summary:
        total, by_type = summary[region]
        bt = ", ".join(f"{k}:{v}" for k, v in sorted(by_type.items()))
        print(f"  {region}: {total} 个实体  ({bt})")
print(f"\n输出: {OUT}/权威实体清单.xlsx  + 权威清单_<region>.csv")
